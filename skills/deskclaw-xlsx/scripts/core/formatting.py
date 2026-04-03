"""Range formatting."""

import json

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.utils.cell import range_boundaries
except ImportError:
    load_workbook = None


def _check_openpyxl():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def _norm_hex(v):
    if v is None:
        return None
    x = str(v).replace('#','').upper()
    if len(x)==6:
        return 'FF'+x
    return x


def format_range(file_path, sheet_name, start_cell, end_cell=None, bold=False, italic=False, underline=False, font_size=None, font_color=None, bg_color=None, border_style=None, border_color=None, number_format=None, alignment=None, wrap_text=False, merge_cells=False, protection=None, conditional_format=None):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    rng = start_cell if not end_cell else f"{start_cell}:{end_cell}"
    min_col, min_row, max_col, max_row = range_boundaries(rng)

    side = None
    if border_style:
        side = Side(style=border_style, color=_norm_hex(border_color) if border_color else None)
    border = Border(left=side, right=side, top=side, bottom=side) if side else None

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = Font(
                name=cell.font.name,
                sz=font_size if font_size else cell.font.sz,
                bold=bold or cell.font.bold,
                italic=italic or cell.font.italic,
                underline='single' if underline else cell.font.underline,
                color=_norm_hex(font_color) if font_color else cell.font.color,
            )
            if bg_color:
                cell.fill = PatternFill(fill_type='solid', fgColor=_norm_hex(bg_color))
            if border:
                cell.border = border
            if number_format:
                cell.number_format = number_format
            if alignment:
                cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
            elif wrap_text:
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal, wrap_text=True)
            if protection:
                cell.protection = Protection(locked=bool(protection.get('locked', True)), hidden=bool(protection.get('hidden', False)))

    if merge_cells and end_cell:
        ws.merge_cells(f"{start_cell}:{end_cell}")

    wb.save(file_path)
    wb.close()
    return "Range formatted successfully"


def validate_range(file_path, sheet_name, start_cell, end_cell=None):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    try:
        rng = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        _ = ws[rng]
    except Exception as e:
        wb.close()
        return f"Error: invalid range {rng}: {e}"
    wb.close()
    return f"Range {rng} is valid"
