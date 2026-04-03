"""Chart creation."""

try:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, BarChart, PieChart, ScatterChart, Reference, Series
    from openpyxl.utils.cell import range_boundaries, coordinate_from_string, column_index_from_string
except ImportError:
    load_workbook = None


def _check_openpyxl():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def _coord(cell):
    col,row = coordinate_from_string(cell)
    return row, column_index_from_string(col)


def create_chart(file_path, sheet_name, data_range, chart_type, target_cell, title='', x_axis='', y_axis=''):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(data_range)
    chart_type = chart_type.lower()
    if chart_type == 'line':
        chart = LineChart()
    elif chart_type in ('bar', 'column'):
        chart = BarChart()
    elif chart_type == 'pie':
        chart = PieChart()
    elif chart_type == 'scatter':
        chart = ScatterChart()
    else:
        wb.close(); return f"Error: unsupported chart_type '{chart_type}'"

    data = Reference(ws, min_col=min_col+1 if max_col>min_col else min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    if max_col > min_col:
        cats = Reference(ws, min_col=min_col, min_row=min_row+1, max_row=max_row)
        chart.set_categories(cats)
    if title:
        chart.title = title
    if x_axis:
        chart.x_axis.title = x_axis
    if y_axis:
        chart.y_axis.title = y_axis
    ws.add_chart(chart, target_cell)
    wb.save(file_path); wb.close()
    return f"Created {chart_type} chart at {target_cell}"
