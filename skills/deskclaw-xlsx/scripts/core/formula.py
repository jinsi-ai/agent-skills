"""Formula apply and validation."""

try:
    from openpyxl import load_workbook
    from openpyxl.formula.tokenizer import Tokenizer
except ImportError:
    load_workbook = None


def _check_openpyxl():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def validate_formula(file_path, sheet_name, cell, formula):
    _check_openpyxl()
    if not str(formula).startswith('='):
        return "Error: formula must start with '='"
    try:
        Tokenizer(formula)
        return f"Formula is valid for {sheet_name}!{cell}"
    except Exception as e:
        return f"Error: invalid formula syntax: {e}"


def apply_formula(file_path, sheet_name, cell, formula):
    _check_openpyxl()
    v = validate_formula(file_path, sheet_name, cell, formula)
    if v.startswith('Error:'):
        return v
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    ws[cell] = formula
    wb.save(file_path)
    wb.close()
    return f"Applied formula to {sheet_name}!{cell}"
