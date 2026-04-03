"""Worksheet and range operations."""

from copy import copy

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError:
    load_workbook = None


def _check_openpyxl():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def _coord(cell):
    col,row = coordinate_from_string(cell)
    return row, column_index_from_string(col)


def copy_worksheet(file_path, source_sheet, target_sheet):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if source_sheet not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{source_sheet}' not found"
    if target_sheet in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{target_sheet}' already exists"
    ws = wb.copy_worksheet(wb[source_sheet])
    ws.title = target_sheet
    wb.save(file_path); wb.close()
    return f"Copied worksheet '{source_sheet}' to '{target_sheet}'"


def delete_worksheet(file_path, sheet_name):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    if len(wb.sheetnames) == 1:
        wb.close(); return "Error: cannot delete the last worksheet"
    del wb[sheet_name]
    wb.save(file_path); wb.close()
    return f"Deleted worksheet '{sheet_name}'"


def rename_worksheet(file_path, old_name, new_name):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if old_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{old_name}' not found"
    if new_name in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{new_name}' already exists"
    wb[old_name].title = new_name
    wb.save(file_path); wb.close()
    return f"Renamed worksheet '{old_name}' to '{new_name}'"


def merge_cells(file_path, sheet_name, start_cell, end_cell):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    ws.merge_cells(f"{start_cell}:{end_cell}")
    wb.save(file_path); wb.close()
    return f"Merged cells {start_cell}:{end_cell}"


def unmerge_cells(file_path, sheet_name, start_cell, end_cell):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    ws.unmerge_cells(f"{start_cell}:{end_cell}")
    wb.save(file_path); wb.close()
    return f"Unmerged cells {start_cell}:{end_cell}"


def get_merged_cells(file_path, sheet_name):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    out = [str(rng) for rng in ws.merged_cells.ranges]
    wb.close()
    return str(out)


def copy_range(file_path, sheet_name, source_start, source_end, target_start, target_sheet=None):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    tgt_name = target_sheet or sheet_name
    if tgt_name not in wb.sheetnames:
        wb.close(); return f"Error: target sheet '{tgt_name}' not found"
    src = wb[sheet_name]
    tgt = wb[tgt_name]
    sr1, sc1 = _coord(source_start)
    sr2, sc2 = _coord(source_end)
    tr1, tc1 = _coord(target_start)
    for r in range(sr1, sr2+1):
        for c in range(sc1, sc2+1):
            src_cell = src.cell(r,c)
            tr = tr1 + (r - sr1)
            tc = tc1 + (c - sc1)
            dst_cell = tgt.cell(tr,tc)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
    wb.save(file_path); wb.close()
    return f"Copied range {source_start}:{source_end} to {tgt_name}!{target_start}"


def delete_range(file_path, sheet_name, start_cell, end_cell, shift_direction='up'):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    r1,c1 = _coord(start_cell)
    r2,c2 = _coord(end_cell)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r,c).value = None
    if shift_direction == 'up':
        ws.move_range(f"{start_cell}:{end_cell}", rows=-(r2-r1+1), cols=0)
    elif shift_direction == 'left':
        ws.move_range(f"{start_cell}:{end_cell}", rows=0, cols=-(c2-c1+1))
    wb.save(file_path); wb.close()
    return f"Deleted range {start_cell}:{end_cell}"


def insert_rows(file_path, sheet_name, start_row, count=1):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    wb[sheet_name].insert_rows(start_row, count)
    wb.save(file_path); wb.close()
    return f"Inserted {count} row(s) at {start_row}"


def insert_columns(file_path, sheet_name, start_col, count=1):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    wb[sheet_name].insert_cols(start_col, count)
    wb.save(file_path); wb.close()
    return f"Inserted {count} column(s) at {start_col}"


def delete_rows(file_path, sheet_name, start_row, count=1):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    wb[sheet_name].delete_rows(start_row, count)
    wb.save(file_path); wb.close()
    return f"Deleted {count} row(s) from {start_row}"


def delete_columns(file_path, sheet_name, start_col, count=1):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    wb[sheet_name].delete_cols(start_col, count)
    wb.save(file_path); wb.close()
    return f"Deleted {count} column(s) from {start_col}"
