"""Excel table and pivot-like summary table creation."""

from collections import defaultdict

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.cell import range_boundaries
except ImportError:
    load_workbook = None


def _check_openpyxl():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def create_table(file_path, sheet_name, data_range, table_name=None, table_style='TableStyleMedium9'):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    if table_name is None:
        table_name = f"Table_{sheet_name.replace(' ', '_')}"
    tab = Table(displayName=table_name, ref=data_range)
    style = TableStyleInfo(name=table_style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(file_path); wb.close()
    return f"Created table '{table_name}' in range {data_range}"


def create_pivot_table(file_path, sheet_name, data_range, rows, values, columns=None, agg_func='mean'):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close(); return f"Error: Sheet '{sheet_name}' not found"
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(data_range)

    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col+1)]
    idx = {str(h): i for i, h in enumerate(headers) if h is not None}
    for key in rows + values + (columns or []):
        if key not in idx:
            wb.close(); return f"Error: column '{key}' not found in header row"

    data_rows = []
    for r in range(min_row+1, max_row+1):
        vals = [ws.cell(r, c).value for c in range(min_col, max_col+1)]
        data_rows.append(vals)

    group = defaultdict(list)
    for rec in data_rows:
        rkey = tuple(rec[idx[k]] for k in rows)
        ckey = tuple(rec[idx[k]] for k in (columns or [])) if columns else tuple()
        v = []
        for name in values:
            try:
                v.append(float(rec[idx[name]]) if rec[idx[name]] is not None else 0.0)
            except Exception:
                v.append(0.0)
        group[(rkey, ckey)].append(v)

    def agg(nums):
        if not nums:
            return 0
        if agg_func == 'sum':
            return sum(nums)
        if agg_func == 'count':
            return len(nums)
        if agg_func == 'max':
            return max(nums)
        if agg_func == 'min':
            return min(nums)
        return sum(nums)/len(nums)

    out_name = f"Pivot_{sheet_name}"
    i = 1
    while out_name in wb.sheetnames:
        i += 1
        out_name = f"Pivot_{sheet_name}_{i}"
    out = wb.create_sheet(out_name)

    col_keys = sorted(set(k[1] for k in group.keys())) if columns else [tuple()]
    header = list(rows)
    for ck in col_keys:
        label = '|'.join(str(x) for x in ck) if ck else 'All'
        for val_name in values:
            header.append(f"{label}:{val_name}")
    out.append(header)

    row_keys = sorted(set(k[0] for k in group.keys()))
    for rk in row_keys:
        row_out = list(rk)
        for ck in col_keys:
            records = group.get((rk, ck), [])
            for i_val in range(len(values)):
                nums = [rec[i_val] for rec in records]
                row_out.append(agg(nums))
        out.append(row_out)

    wb.save(file_path); wb.close()
    return f"Created pivot-like summary sheet '{out_name}'"
