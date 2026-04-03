"""Read and write worksheet data."""

import json

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError:
    load_workbook = None


def _check_openpyxl():
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def _coord(cell):
    col,row = coordinate_from_string(cell)
    return row, column_index_from_string(col)


def read_data(file_path, sheet_name, start_cell='A1', end_cell=None, preview_only=False):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return json.dumps({"error": f"Sheet '{sheet_name}' not found"})
    ws = wb[sheet_name]
    if end_cell is None:
        end_cell = ws.dimensions.split(':')[-1]
    sr, sc = _coord(start_cell)
    er, ec = _coord(end_cell)
    rows=[]
    for r in range(sr, er+1):
        row=[]
        for c in range(sc, ec+1):
            row.append(ws.cell(r,c).value)
        rows.append(row)
    wb.close()
    if preview_only:
        rows = rows[:20]
    return json.dumps({"sheet_name": sheet_name, "range": f"{start_cell}:{end_cell}", "rows": rows}, ensure_ascii=False, indent=2, default=str)


def write_data(file_path, sheet_name, data, start_cell='A1'):
    _check_openpyxl()
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    sr, sc = _coord(start_cell)
    for i, row in enumerate(data):
        for j, val in enumerate(row):
            ws.cell(sr+i, sc+j).value = val
    wb.save(file_path)
    wb.close()
    return f"Wrote {len(data)} row(s) to {sheet_name} at {start_cell}"
