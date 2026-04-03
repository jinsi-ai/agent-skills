"""Workbook operations."""

import json
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


def _check_openpyxl():
    if Workbook is None or load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")


def create_workbook(file_path):
    _check_openpyxl()
    wb = Workbook()
    wb.save(file_path)
    return f"Created workbook at {file_path}"


def create_worksheet(file_path, sheet_name):
    _check_openpyxl()
    wb = load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        wb.close()
        return f"Error: sheet '{sheet_name}' already exists"
    wb.create_sheet(sheet_name)
    wb.save(file_path)
    wb.close()
    return f"Created worksheet: {sheet_name}"


def get_workbook_metadata(file_path, include_ranges=False):
    _check_openpyxl()
    wb = load_workbook(file_path)
    sheets = []
    for s in wb.worksheets:
        entry = {
            "name": s.title,
            "max_row": s.max_row,
            "max_col": s.max_column,
            "dimensions": s.dimensions,
        }
        if include_ranges:
            entry["merged_ranges"] = [str(rng) for rng in s.merged_cells.ranges]
        sheets.append(entry)
    out = {
        "file_path": str(file_path),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "sheets": sheets,
    }
    wb.close()
    return json.dumps(out, ensure_ascii=False, indent=2)
